"""Excel spreadsheet (.xlsx) to CSV/Markdown converter."""

from io import BytesIO

from openpyxl import load_workbook

from .base import ConversionError, ConversionResult, strip_extension


def convert_xlsx(file_bytes: bytes, filename: str) -> ConversionResult:
    """Convert an Excel spreadsheet to Markdown tables.

    For multi-sheet workbooks, each sheet becomes a separate section.
    Empty sheets are skipped.

    Args:
        file_bytes: Raw .xlsx file content
        filename: Original filename

    Returns:
        ConversionResult with Markdown content (tables per sheet)
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ConversionError(f"Failed to parse Excel file: {e}") from e

    markdown_sections: list[str] = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        md_table = _sheet_to_markdown(sheet, sheet_name)
        # Always include sheet - note if empty to preserve information
        if md_table:
            markdown_sections.append(md_table)
        else:
            markdown_sections.append(f"## {sheet_name}\n\n*Empty sheet*")

    wb.close()

    if not markdown_sections:
        # Empty workbook - return minimal content
        markdown_content = f"# {strip_extension(filename)}\n\n*Empty spreadsheet*"
    else:
        markdown_content = "\n\n".join(markdown_sections)

    new_filename = f"{strip_extension(filename)}.md"

    return ConversionResult(
        content=markdown_content.encode("utf-8"),
        mime_type="text/markdown",
        filename=new_filename,
    )


def _sheet_to_markdown(sheet, sheet_name: str) -> str | None:
    """Convert a worksheet to a Markdown table."""
    rows: list[list[str]] = []

    for row in sheet.iter_rows():
        cell_values = []
        for cell in row:
            value = cell.value
            if value is None:
                cell_values.append("")
            else:
                # Convert to string and clean up
                text = str(value).strip().replace("\n", " ").replace("|", "\\|")
                cell_values.append(text)
        rows.append(cell_values)

    # Remove trailing empty rows
    while rows and all(cell == "" for cell in rows[-1]):
        rows.pop()

    # Remove leading empty rows
    while rows and all(cell == "" for cell in rows[0]):
        rows.pop(0)

    if not rows:
        return None

    # Find actual column count (ignore trailing empty columns)
    max_cols = 0
    for row in rows:
        for i, cell in enumerate(reversed(row)):
            if cell:
                max_cols = max(max_cols, len(row) - i)
                break

    if max_cols == 0:
        return None

    # Trim rows to max_cols
    rows = [row[:max_cols] for row in rows]

    # Pad short rows
    for row in rows:
        while len(row) < max_cols:
            row.append("")

    # Build markdown
    lines: list[str] = []

    # Sheet name as heading
    lines.append(f"## {sheet_name}")
    lines.append("")

    # Header row
    header = rows[0]
    lines.append("| " + " | ".join(header) + " |")

    # Separator
    lines.append("| " + " | ".join(["---"] * len(header)) + " |")

    # Data rows
    for row in rows[1:]:
        lines.append("| " + " | ".join(row) + " |")

    return "\n".join(lines)
